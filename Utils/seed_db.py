# This file contains code used to read the excel sheet containing student info
# and inserts it into the database

from openpyxl import load_workbook
import mysql.connector
from mysql.connector import Error

from dateutil.parser import parse

# helper function to transform "Not Applicable" to 0
def transform_na(student_info: list, index: int):
    if student_info[index] == "Not Applicable":
        student_info[index] = 0
    else:
        student_info[index] = 1


def get_maximum_data_rows(sheet_object):
    '''Gets the max number of rows in the excel worksheet.
    These rows contain actual data instead of being empty'''
    rows = 0
    for _, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows

# read excel sheet
path_to_excel = input("Enter the path to your excel data file: ")

worksheet = load_workbook(path_to_excel, read_only=True).worksheets[0]
    

# connect to database
try:
    db_host = input("Enter database host (e.g localhost): ")
    db_name = input("Enter database name (e.g. mydb): ")
    db_user = input("Enter database user name (e.g. root): ")
    db_password = input("Enter database password: ")

    connection = mysql.connector.connect(host=db_host,
                                        database=db_name,
                                        user=db_user,
                                        password=db_password)

    # we will commit all insertions after every single one is done
    connection.autocommit = False

    cursor = connection.cursor()
    
    student_auth_insert_query = """INSERT INTO student_auth(prn, password, payment_done, antiragging_uploaded)
                                    VALUES (%s, %s, %s, %s)"""

    student_data_insert_query = """INSERT INTO student_data(application_id, prn, name, category, gender, fathers_name,
                                    mothers_name, date_of_birth, religion, region, mother_tongue, annual_income_start,
                                    annual_income_end, line_1, line_2, line_3, state, district, taluka, village, pincode, mobile_number,
                                    email_address, physically_handicapped, linguistic_minority, religious_minority, ssc_board, ssc_passing_year,
                                    ssc_total_percentage, qualifying_exam, hsc_board, hsc_passing_year, physics_percentage, chemistry_percentage,
                                    math_percentage, hsc_total_percentage, eligibility_percentage, cet_percentile, jee_percentile, merit_no, merit_marks,
                                    course_name, admission_date)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

    # get the maximum number of rows containing actual data
    max_rows = get_maximum_data_rows(worksheet)

    for row in worksheet.iter_rows(2, max_rows):
        row_list = list(row)
        
        student_prn = row_list[2].value
        
        student_data = []

        for i in row_list[1:12]:
            student_data.append(i.value)

        
        (income_start, income_end) = row_list[12].value.split(" - ")

        income_start = int(income_start.replace(",", ""))
        income_end = int(income_end.replace(",", ""))

        student_data.append(income_start)
        student_data.append(income_end)

        student_data.extend([x.value for x in row_list[13:]])

        student_data[20] = int(student_data[20])
        student_data[21] = int(student_data[21])
        student_data[27] = int(student_data[27])
        student_data[31] = int(student_data[31])
        student_data[39] = int(student_data[39])

        # change "Not Applicable values to 1 or 0"
        transform_na(student_data, 23)
        transform_na(student_data, 24)
        transform_na(student_data, 25)

        # Convert the birth date to password
        birthdate = student_data[7]
        password = f"{birthdate.day}{birthdate.month}{birthdate.year}"

        # record to insert in student_auth table
        student_auth = [student_prn, password, 0, 0]

        # insert student_auth
        cursor.execute(student_auth_insert_query, student_auth)

        # insert student_data
        cursor.execute(student_data_insert_query, student_data)

    connection.commit()
    print("Data inserted successfully.")

except Error as e:
    print("Failed to update record to database rollback: {}".format(e))

    # reverting changes because of exception
    connection.rollback()
finally:
    # closing database connection.
    if connection.is_connected():
        cursor.close()
        connection.close()
        print("connection is closed")
